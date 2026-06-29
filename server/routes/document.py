import os
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename
from extensions import db
from models.document import Document
from config import Config

document_bp = Blueprint("document", __name__)

ALLOWED_EXTENSIONS = {"pdf", "docx", "txt", "md", "csv", "xlsx"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_type(filename):
    if "." in filename:
        return filename.rsplit(".", 1)[1].lower()
    return "unknown"

def read_file_content(file_path, file_type):
    content = None
    if file_type in ("txt", "md", "csv"):
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
    elif file_type == "pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text: parts.append(text)
            content = "\n".join(parts) if parts else None
        except Exception:
            content = None
    elif file_type == "docx":
        try:
            from docx import Document as DocxReader
            doc = DocxReader(file_path)
            content = "\n".join([p.text for p in doc.paragraphs])
        except Exception:
            content = None
    elif file_type == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append("Sheet: " + sheet_name)
                for row in ws.iter_rows(values_only=True):
                    line = " | ".join([str(c) if c is not None else "" for c in row])
                    lines.append(line)
            content = "\n".join(lines)
        except Exception:
            content = None
    return content

@document_bp.route("/upload", methods=["POST"])
@jwt_required()
def upload_document():
    user_id = int(get_jwt_identity())

    if "file" not in request.files:
        return jsonify({"code": 400, "msg": "No file uploaded"}), 200

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"code": 400, "msg": "No file selected"}), 200

    original_filename = file.filename
    if not allowed_file(original_filename):
        return jsonify({"code": 400, "msg": "File type not allowed"}), 200

    filename = secure_filename(original_filename)
    if not filename or "." not in filename:
        import uuid
        ext = get_file_type(original_filename)
        filename = f"{uuid.uuid4().hex}.{ext}"

    upload_folder = os.path.join(Config.UPLOAD_FOLDER, str(user_id))
    os.makedirs(upload_folder, exist_ok=True)
    file_path = os.path.join(upload_folder, filename)
    file.save(file_path)

    file_type = get_file_type(original_filename)

    doc = Document(
        filename=original_filename,
        file_path=file_path,
        file_type=file_type,
        file_size=os.path.getsize(file_path),
        user_id=user_id
    )
    db.session.add(doc)
    db.session.commit()

    content = read_file_content(file_path, file_type)
    msg = "Upload successful"
    if content:
        try:
            from services.embedding_service import EmbeddingService
            es = EmbeddingService()
            ok = es.split_and_vectorize(doc.id, original_filename, content)
            if not ok:
                msg = "Uploaded but vectorization failed"
        except Exception as e:
            msg = f"Uploaded but vectorization error: {e}"
    else:
        msg = "Uploaded but could not extract text"

    return jsonify({"code": 200, "msg": msg, "data": doc.to_dict()})

@document_bp.route("/list", methods=["GET"])
@jwt_required()
def list_documents():
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    query = Document.query.filter_by(status=1)
    total = query.count()
    docs = query.order_by(Document.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return jsonify({"code": 200, "data": {"items": [d.to_dict() for d in docs], "total": total}})

@document_bp.route("/<int:doc_id>", methods=["DELETE"])
@jwt_required()
def delete_document(doc_id):
    doc = Document.query.filter_by(id=doc_id).first()
    if not doc:
        return jsonify({"code": 404, "msg": "Document not found"}), 200
    doc.status = 0
    db.session.commit()
    return jsonify({"code": 200, "msg": "Delete successful"})